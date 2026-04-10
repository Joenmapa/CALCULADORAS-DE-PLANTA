"""
RecipeManager - Servidor Flask
Lee el Excel de la carpeta compartida y sirve los datos como API REST.

Instalación:
    pip install flask openpyxl flask-cors

Uso:
    python server.py
"""

import json
import os
from flask import Flask, jsonify
from flask_cors import CORS
import openpyxl

app = Flask(__name__)
CORS(app)  # permite que el HTML en cualquier PC de la red pueda hacer fetch()

# ── Configuración ────────────────────────────────────────────────────────────
CONFIG_FILE = os.path.join(os.path.dirname(__file__), "excel_config.json")

def load_config():
    with open(CONFIG_FILE, "r", encoding="utf-8") as f:
        return json.load(f)

# ── Lector del Excel ─────────────────────────────────────────────────────────
def read_excel():
    config = load_config()
    excel_path = config["excel_path"]

    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"No se encontró el Excel en: {excel_path}")

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    sheets = config["sheets"]

    # ── Hoja: Productos ──────────────────────────────────────────────────────
    products = []
    ws_prod = wb[sheets["products"]]
    rows_prod = list(ws_prod.iter_rows(values_only=True))
    headers_prod = [str(h).strip() if h else "" for h in rows_prod[0]]

    col = {name: i for i, name in enumerate(headers_prod)}
    mapping_prod = config["columns"]["products"]

    for row in rows_prod[1:]:
        if not row[col.get(mapping_prod["code"], 0)]:
            continue  # saltar filas vacías

        product = {
            "code":        str(row[col[mapping_prod["code"]]]).strip(),
            "description": str(row[col[mapping_prod["description"]]]).strip() if row[col[mapping_prod["description"]]] else "",
            "brand":       str(row[col[mapping_prod.get("brand", "")]]).strip() if mapping_prod.get("brand") and row[col.get(mapping_prod.get("brand",""), 0)] else "",
            "flavor":      str(row[col[mapping_prod.get("flavor", "")]]).strip() if mapping_prod.get("flavor") and row[col.get(mapping_prod.get("flavor",""), 0)] else "",
            "dissolution": row[col[mapping_prod["dissolution"]]] if mapping_prod.get("dissolution") and col.get(mapping_prod.get("dissolution","")) is not None else 1,
            "batch_size":  float(row[col[mapping_prod["batch_size"]]]) if mapping_prod.get("batch_size") and row[col.get(mapping_prod.get("batch_size",""), 0)] else 0,
            "line":        str(row[col[mapping_prod.get("line", "")]]).strip() if mapping_prod.get("line") and row[col.get(mapping_prod.get("line",""), 0)] else "",
            "recipe":      {},
            "attrs":       {},
        }
        products.append(product)

    # ── Hoja: Materias Primas ────────────────────────────────────────────────
    raw_materials = []
    ws_rm = wb[sheets["raw_materials"]]
    rows_rm = list(ws_rm.iter_rows(values_only=True))
    headers_rm = [str(h).strip() if h else "" for h in rows_rm[0]]
    col_rm = {name: i for i, name in enumerate(headers_rm)}
    mapping_rm = config["columns"]["raw_materials"]

    for row in rows_rm[1:]:
        if not row[col_rm.get(mapping_rm["code"], 0)]:
            continue
        rm = {
            "code":        str(row[col_rm[mapping_rm["code"]]]).strip(),
            "description": str(row[col_rm[mapping_rm["description"]]]).strip() if row[col_rm[mapping_rm["description"]]] else "",
            "material":    str(row[col_rm[mapping_rm.get("material", "")]]).strip() if mapping_rm.get("material") and row[col_rm.get(mapping_rm.get("material",""), 0)] else "",
            "unit":        str(row[col_rm[mapping_rm["unit"]]]).strip() if row[col_rm[mapping_rm["unit"]]] else "",
            "brix":        float(row[col_rm[mapping_rm["brix"]]]) if mapping_rm.get("brix") and row[col_rm.get(mapping_rm.get("brix",""), 0)] else 0,
        }
        raw_materials.append(rm)

    # Índice de materias primas para lookup rápido
    rm_map = {rm["code"]: rm for rm in raw_materials}

    # ── Hoja: Recetas ────────────────────────────────────────────────────────
    # Formato esperado: columnas = ProductCode | IngredientCode | Quantity
    if sheets.get("recipes"):
        ws_rec = wb[sheets["recipes"]]
        rows_rec = list(ws_rec.iter_rows(values_only=True))
        headers_rec = [str(h).strip() if h else "" for h in rows_rec[0]]
        col_rec = {name: i for i, name in enumerate(headers_rec)}
        mapping_rec = config["columns"]["recipes"]

        prod_map = {p["code"]: p for p in products}

        for row in rows_rec[1:]:
            prod_code = row[col_rec.get(mapping_rec["product_code"], 0)]
            ing_code  = row[col_rec.get(mapping_rec["ingredient_code"], 0)]
            qty       = row[col_rec.get(mapping_rec["quantity"], 0)]
            if not prod_code or not ing_code:
                continue
            prod_code = str(prod_code).strip()
            ing_code  = str(ing_code).strip()
            if prod_code in prod_map:
                prod_map[prod_code]["recipe"][ing_code] = float(qty) if qty else 0

    wb.close()

    return {
        "products":     products,
        "rawMaterials": raw_materials,
        "hoursData":    {},
    }

# ── Endpoints ────────────────────────────────────────────────────────────────
@app.route("/api/data")
def get_data():
    """Devuelve todos los datos frescos del Excel."""
    try:
        data = read_excel()
        return jsonify(data)
    except FileNotFoundError as e:
        return jsonify({"error": str(e)}), 404
    except Exception as e:
        return jsonify({"error": f"Error leyendo Excel: {str(e)}"}), 500
# ── /api/items ───────────────────────────────────────────────
@app.route("/api/items")
def get_items():
    """Solo los productos/items del Excel."""
    try:
        data = read_excel()
        return jsonify(data["products"])
    except FileNotFoundError as e:
        return jsonify({"error": str(e)}), 404
    except Exception as e:
        return jsonify({"error": f"Error leyendo Excel: {str(e)}"}), 500

# ── /api/raw-materials ────────────────────────────────────────
@app.route("/api/raw-materials")
def get_raw_materials():
    """Solo las materias primas del Excel."""
    try:
        data = read_excel()
        return jsonify(data["rawMaterials"])
    except FileNotFoundError as e:
        return jsonify({"error": str(e)}), 404
    except Exception as e:
        return jsonify({"error": f"Error leyendo Excel: {str(e)}"}), 500

# ── /api/status ───────────────────────────────────────────────
@app.route("/api/status")
def get_status():
    """Alias de /api/health para compatibilidad."""
    return health()
@app.route("/api/health")
def health():
    """Endpoint de diagnóstico — abre en el navegador para verificar que el servidor corre."""
    try:
        config = load_config()
        excel_ok = os.path.exists(config["excel_path"])
        return jsonify({
            "status": "ok",
            "excel_found": excel_ok,
            "excel_path": config["excel_path"],
        })
    except Exception as e:
        return jsonify({"status": "error", "detail": str(e)}), 500

# ── Main ─────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("=" * 55)
    print("  RecipeManager API Server")
    print("  Abre en el navegador: http://localhost:5000/api/health")
    print("=" * 55)
    # host="0.0.0.0" hace que sea accesible desde otras PCs de la red
    app.run(host="0.0.0.0", port=5000, debug=False)
